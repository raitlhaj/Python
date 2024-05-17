#scrapping
import pandas as pd
import openpyxl as xl
from openpyxl.chart import BarChart,Reference
from pathlib import Path
import requests
from datetime import datetime, timedelta
import investpy


def yesterday(frmt='%m/%d/%Y', string=True):
    yesterday = datetime.now() - timedelta(1)
    if string:
        return yesterday.strftime(frmt)
    return yesterday

def process_workoob(filename):
    
    path=Path(filename)
  
    print(yesterday)
    if(path.exists()):
        print('the file has been found !')
    else:
        print('the file not found !')

        
    wb=xl.load_workbook(filename)
    sh=wb['Stock']
    dicto=investpy.stocks.get_stocks(country=None)
 
    for col in range(1,sh.max_column+1):
        
        ticker= str(sh.cell(1,col).value)

        try:
             country='united states'
             url = "http://api.scraperlink.com/investpy/?email=ait_lhaj92@live.fr&type=historical_data&product=stocks&country="+str(country)+"&symbol="+str(ticker)+"&from_date=01/01/2019&to_date="+yesterday()
             response = requests.get(url)
             resp_dict = response.json()
        except:
            
            try:
                country=dicto[dicto['symbol']==ticker].iloc[0]['country']
                url = "http://api.scraperlink.com/investpy/?email=ait_lhaj92@live.fr&type=historical_data&product=stocks&country="+str(country)+"&symbol="+str(ticker)+"&from_date=01/01/2019&to_date="+yesterday()
                response = requests.get(url)
                resp_dict = response.json()
                
            except: 
                print(country+ ": "+ticker+ "=> KO")
                continue
        
        print(country+ ": "+ticker+ "=> OK")
        df = pd.DataFrame(resp_dict.get('data'))

        for row in range(2,len(df['last_close'])):
          sh.cell(row,col).value = df['last_close'][row-1]

    #for row in range(2,sh.max_row+1):
        
        #sh.cell(row,5).value = df['last_close'][0]
        #print(df['last_close'][0])
        



    wb.save(filename)
    
    
process_workoob(r"C:\\Users\Administrateur\Dropbox\Working\Invest\stocksData.xlsx")