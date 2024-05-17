#scrapping
import pandas as pd
import openpyxl as xl
from openpyxl.chart import BarChart,Reference
from pathlib import Path
import requests
from datetime import datetime, timedelta
import investpy
from random import *
import openpyxl as xl

def yesterday(frmt='%m/%d/%Y', string=True):
    yesterday = datetime.now() - timedelta(1)
    if string:
        return yesterday.strftime(frmt)
    return yesterday

def export_quotes_to_excel(start_date: str, tickers : list[str]):
    chaine = ""
    col = 1
    dicto = investpy.stocks.get_stocks(country=None)
    wb=xl.Workbook()
    sh=wb.active
    
    for ticker in tickers:
        
        col = col +1
        sh.cell(1,col).value = ticker
        
        try:
             country='united states'
             url = "http://api.scraperlink.com/investpy/?email=ait_lhaj92@live.fr&type=historical_data&product=stocks&country="+str(country)+"&symbol="+str(ticker)+"&from_date="+str(start_date)+"&to_date="+yesterday()
             response = requests.get(url)
             resp_dict = response.json()
        except:
            
            try:
                country=dicto[dicto['symbol']==ticker].iloc[0]['country']
                url = "http://api.scraperlink.com/investpy/?email=ait_lhaj92@live.fr&type=historical_data&product=stocks&country="+str(country)+"&symbol="+str(ticker)+"&from_date="+str(start_date)+"&to_date="+yesterday()
                response = requests.get(url)
                resp_dict = response.json()              
            except: 
                chaine = str(chaine) + country + ": "+ticker+ "=> KO \n"
                continue
        
        chaine = str(chaine) + country + ": "+ ticker + "=> OK \n"
        df = pd.DataFrame(resp_dict.get('data'))

        for row in range(2,len(df['last_close'])):
          sh.cell(row,col).value = df['last_close'][row-1]

        try:
            wb.save(r"C:\\Users\Administrateur\Dropbox\Working\Invest\export_quotes_to_excel.xlsx")
        except:
            print("Operation failed !")
            
    print(chaine+ "check export_quotes_to_excel.xlsx" )
    
