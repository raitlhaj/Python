import openpyxl as xl
from openpyxl.chart import BarChart,Reference
from pathlib import Path

#from paylrowb import path


def process_workoob(filename):
    
    path=Path(filename)
    if(path.exists()):
        print('the file has been found !')
    else:
        print('the file not found !')

        
    wb=xl.load_workbook(filename)
    sh=wb['Feuil1']
    #2 ways to accces to the cell
    cel=sh['A1']
    cel=sh.cell(1,1)

    for row in range(2,sh.max_row+1):
        sh.cell(row,5).value=int(sh.cell(row,2).value)*int(sh.cell(row,3).value)
        
    values=Reference(sh, 
            min_row=2,
            max_row=sh.max_row,
            min_col=5,
            max_col=5)
    chart=BarChart()
    chart.add_data(values)
    sh.add_chart(chart,'f4')


    wb.save("products.xlsx")
    
    
process_workoob("products.xlsx")